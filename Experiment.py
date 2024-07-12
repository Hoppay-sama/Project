
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox, ttk
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl import load_workbook

exl = Workbook()
exl = load_workbook('AMS.xlsx')

cmd = tk.Tk()

#computations
heights= 800
widths= 1200
screenx= cmd.winfo_screenwidth()
screeny= cmd.winfo_screenheight()
leftscrn= int(screenx/2 - widths/2)
topscrn= 0
topscrn2= int(screeny/2 - (heights-200)/2)

cmd.geometry(f'{widths}x{heights-200}+{leftscrn}+{topscrn2}')
cmd.title('Apartment Manager')
cmd.iconbitmap('TetraICO.ico')

#Images
img1= Image.open('Preview1.png')
Animation= ImageTk.PhotoImage(img1)
img2 = Image.open('Logout.png')
Logout= ImageTk.PhotoImage(img2)
door1 = Image.open('Vacant.png')
Vacant= ImageTk.PhotoImage(door1)
door2 = Image.open('Occupied.png')
Occupied= ImageTk.PhotoImage(door2)
profile2 = Image.open('profile.png')
Profile1= ImageTk.PhotoImage(profile2)
pn = Image.open('companion.png')
pnp= ImageTk.PhotoImage(pn)
oth = Image.open('self.png')
self= ImageTk.PhotoImage(oth)


#Functions
def tab1(): #Home
  Home.configure(state='disabled', cursor='')
  Set.grid_columnconfigure(0, weight=1)
  Set.grid_rowconfigure(0,weight=1)
  Set.grid(row=1, column=0, sticky='nsew')
  Preview.pack()
  Options.pack()
  Room.configure(state='normal', cursor='hand2')
  Building.grid_forget()
  Employee.configure(state='normal', cursor='hand2')
  Frame1.grid_forget()
  Containter.pack_forget()
  reset()
  ascend()

def tab2(): #Apartment
  cmd.after_cancel(loop1)
  global Home, Room
  Home.configure(state='normal', cursor='hand2')
  Preview.pack_forget()
  Options.pack_forget()
  Set.grid_forget()  
  Room.configure(state='disabled', cursor='')
  Building.grid_columnconfigure(0, weight=1)
  Building.grid_rowconfigure(0,weight=1)
  Building.grid(row=1, column=0, sticky='nsew')
  Floor.pack(side = 'left', fill='both')
  Label(Building, text='Room 101', width=15, font=('System',17), bg='#FFF59D').place(x=100, y=60)
  Label(Building, text='Room 102', width=15, font=('System',17), bg='#FFF59D').place(x=460, y=60)
  Label(Building, text='Room 103', width=15, font=('System',17), bg='#FFF59D').place(x=820, y=60)
  Label(Building, text='Room 104', width=15, font=('System',17), bg='#FFF59D').place(x=100, y=460)
  Label(Building, text='Room 105', width=15, font=('System',17), bg='#FFF59D').place(x=460, y=460)
  Label(Building, text='Room 106', width=15, font=('System',17), bg='#FFF59D').place(x=820, y=460)
  A1()
  A2()
  A3()
  A4()
  A5()
  A6()
  Employee.configure(state='normal', cursor='hand2')
  Frame1.grid_forget()
  Containter.pack_forget()

def tab5(): #Employee
  cmd.after_cancel(loop1)
  global Home, Room
  Home.configure(state='normal', cursor='hand2')
  Preview.pack_forget()
  Options.pack_forget()
  Set.grid_forget()
  Room.configure(state='normal', cursor='hand2')
  Building.grid_forget()
  Employee.configure(state='disabled', cursor='')
  #new
  Frame1.grid_columnconfigure(0, weight=1)
  Frame1.grid_rowconfigure(0,weight=1)
  Frame1.grid(row=1, column=0, sticky='nsew')
  Containter.pack(side = 'left', fill='both')
  ctk.CTkLabel(Containter, text='Employee Name:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=30, y=50)
  ctk.CTkLabel(Containter, text='Occupation:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=550, y=50)
  ctk.CTkLabel(Containter, text='Days of Work:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=30, y=150)
  ctk.CTkLabel(Containter, text='Work hours:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=550, y=150)
  ctk.CTkLabel(Containter, text='Monthly Salary:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=30, y=250)
  ctk.CTkLabel(Containter, text='Age:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=550, y=250)
  ctk.CTkLabel(Containter, text='Date of Birth:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=30, y=350)
  ctk.CTkLabel(Containter, text='Sex:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=550, y=350)
  ctk.CTkLabel(Containter, text='Note:', fg_color='transparent', font=('Engravers MT',20), text_color='black').place(x=30, y=450)
  Employ()

#animations
def ascend():
  global moving, loop1
  if moving != -700:
    moving -= 2
    anim.place(x=0, y=moving)
    loop1= cmd.after(40,ascend)
  else:
    loop1= cmd.after(40,descend)
def descend():
  global moving, loop1
  if moving != 0:
    moving += 2
    anim.place(x=0, y=moving)
    loop1= cmd.after(40,descend)
  else:
    loop1= cmd.after(40,ascend)

def run(func):
  global main, domain, constant, side1, lft, rght
  cmd.after_cancel(right)
  cmd.after_cancel(left)
  Main.lower()
  domain= Main.winfo_x()
  func()
def left():
  global main, domain, constant, side1, lft
  if main > -990:
    Left.configure(state='disabled')
    Right.configure(state='normal')
    main -= 10
    side1 -= 10
    Main.place(x=main, y= 60)
    Slider1.place(x=side1, y=60)
    lft= cmd.after(1,left)
def right():
  global main, domain, constant, side1, rght
  if main < 110:
    main += 10
    side1 += 10
    Left.configure(state='normal')
    Right.configure(state='disabled')
    Main.place(x=main, y= 60) 
    Slider1.place(x=side1, y=60)
    rght=cmd.after(1,right)

round = True
moving=0
#Home
Set= ctk.CTkScrollableFrame(cmd, height=535, width=1180, fg_color='white', border_width=0, corner_radius=0)
Preview= Frame(Set, width=1200, height=450, bg='blue')
anim= Label(Preview, image=Animation, width=1200, height=1150)
Options= Frame(Set, width=1200, bg='white', height=600)

Main = Frame(Options, width=980, height=450, bg='#C0AA40')
Slider1 = Frame(Options, width=980, height=450, bg='#C0AA40')
Left= ctk.CTkButton(Options, width=100, height=470, fg_color='transparent', text='<', font=('Helvetica', 50), hover_color='#796B28',cursor='hand2', command=lambda:run(left))
Right= ctk.CTkButton(Options, width=100, height=470, fg_color='transparent', text='>', state='disabled', font=('Helvetica', 50), hover_color='#796B28', cursor='hand2', command=lambda:run(right))

Label(Main, text='Your \nApartment \nManagement \nCompanion', font=('System',30), bg='#C0AA40').place(x=100, y=120)
Label(Main, image=pnp, font=('System',30), bg='#C0AA40').place(x=500, y=120)

Label(Slider1, text="Manage \nWho \nYou're \nWorking \nwith", font=('System',30), bg='#C0AA40').place(x=100, y=120)
Label(Slider1, image=self, font=('System',30), bg='#C0AA40').place(x=500, y=120)

Left.place(x=0, y=50)
Right.place(x=1100, y=50)

#Sliders
main = 110
side1 = 1210
constant = 200
Slider1.place(x=side1, y=60)
Main.place(x=main, y= 60)

#Room
def apply(x):
  ws = exl['Tenants']
  if (T1.get() != '') and (T2.get() != '') and (T3.get() != '') and (T4.get() != '') and (T5.get() != '') and (T6.get() != '') and (T7.get() != ''):
    ws['A'+ str(x)]= T1.get()
    ws['B'+ str(x)]= T2.get()
    ws['C'+ str(x)]= T3.get()
    ws['D'+ str(x)]= T4.get()
    ws['E'+ str(x)]= T5.get()
    ws['F'+ str(x)]= T6.get()
    ws['G'+ str(x)]= T7.get()
    ws['H'+ str(x)]= T8.get()
    ws['I'+ str(x)]= T9.get()
    ws['J'+ str(x)]= T10.get()
    ws['K'+ str(x)]= T11.get()
    exl.save('AMS.xlsx')
    rooms.destroy()
    Room_change('Occupied','Red', Occupied, x)
  else:
    messagebox.showerror('Error', "Please Fill the First 7 Entries with Information")

def kick(x):
  ws = exl['Tenants']
  ws['A'+ str(x)]= None
  ws['B'+ str(x)]= None
  ws['C'+ str(x)]= None
  ws['D'+ str(x)]= None
  ws['E'+ str(x)]= None
  ws['F'+ str(x)]= None
  ws['G'+ str(x)]= None
  ws['H'+ str(x)]= None   
  ws['I'+ str(x)]= None
  ws['J'+ str(x)]= None
  ws['K'+ str(x)]= None
  exl.save('AMS.xlsx')
  rooms.destroy()
  Room_change('Vacant','black', Vacant, x)


def A1():
  ws = exl['Tenants']
  if (ws['A2'].value) == None:
    Room1.configure(text='Vacant', activeforeground='black', fg='black', image=Vacant)
    Room1.place(x=100, y=100 )
  elif (ws['A2'].value) != None:
    Room1.configure(text='Occupied', activeforeground='Red', fg='Red', image=Occupied)
    Room1.place(x=100, y=100 )

def A2():
  ws = exl['Tenants']
  if (ws['A3'].value) == None:
    Room2.configure(text='Vacant', activeforeground='black', fg='black', image=Vacant)
    Room2.place(x=460, y=100 )
  elif (ws['A3'].value) != None:
    Room2.configure(text='Occupied', activeforeground='Red', fg='Red', image=Occupied)
    Room2.place(x=460, y=100 )

def A3():
  ws = exl['Tenants']
  if (ws['A4'].value) == None:
    Room3.configure(text='Vacant', activeforeground='black', fg='black', image=Vacant)
    Room3.place(x=820, y=100 )
  elif (ws['A4'].value) != None:
    Room3.configure(text='Occupied', activeforeground='Red', fg='Red', image=Occupied)
    Room3.place(x=820, y=100 )

def A4():
  ws = exl['Tenants']
  if (ws['A5'].value) == None:
    Room4.configure(text='Vacant', activeforeground='black', fg='black', image=Vacant)
    Room4.place(x=100, y=500 )
  elif (ws['A5'].value) != None:
    Room4.configure(text='Occupied', activeforeground='Red', fg='Red', image=Occupied)
    Room4.place(x=100, y=500 )

def A5():
  ws = exl['Tenants']
  if (ws['A6'].value) == None:
    Room5.configure(text='Vacant', activeforeground='black', fg='black', image=Vacant)
    Room5.place(x=460, y=500 )
  elif (ws['A6'].value) != None:
    Room5.configure(text='Occupied', activeforeground='Red', fg='Red', image=Occupied)
    Room5.place(x=460, y=500 )

def A6():
  ws = exl['Tenants']
  if (ws['A7'].value) == None:
    Room6.configure(text='Vacant', activeforeground='black', fg='black', image=Vacant)
    Room6.place(x=820, y=500 )
  elif (ws['A7'].value) != None:
    Room6.configure(text='Occupied', activeforeground='Red', fg='Red', image=Occupied)
    Room6.place(x=820, y=500 )

def Room_change(a,b,c,change):
  if change == 2:
    Room1.configure(text=a, activeforeground=b, fg=b, image=c)
    Room1.place(x=100, y=100 )
  elif change == 3:
    Room2.configure(text=a, activeforeground=b, fg=b, image=c)
    Room2.place(x=460, y=100 )
  elif change == 4:
    Room3.configure(text=a, activeforeground=b, fg=b, image=c)
    Room3.place(x=820, y=100 )
  elif change == 5:
    Room4.configure(text=a, activeforeground=b, fg=b, image=c)
    Room4.place(x=100, y=500 )
  elif change == 6:
    Room5.configure(text=a, activeforeground=b, fg=b, image=c)
    Room5.place(x=460, y=500 )
  elif change == 7:
    Room6.configure(text=a, activeforeground=b, fg=b, image=c)
    Room6.place(x=820, y=500 )  

def occupy(x):
  global T1, T2, T3, T4, T5, T6, T7, T8, T9, T10, T11, rooms
  ws = exl['Tenants']
  x += 1 
  heights= 800
  widths= 600
  screenx= cmd.winfo_screenwidth()
  screeny= cmd.winfo_screenheight()
  leftscrn= int(screenx/2 - widths/2)
  topscrn= int(screeny/2 - (heights)/2)
  rooms = tk.Toplevel()
  rooms.geometry(f'{widths}x{heights}+{leftscrn}+{topscrn}')
  rooms.configure(bg='#FFF9C4')
  rooms.resizable(False, False)
  Profile = Label(rooms, image=Profile1, bg='#FFF9C4')
  Profile.place(x=4, y=2)
  if ws['A'+str(x)].value != None:
    Label(rooms, text=str(ws['A'+str(x)].value), bg='#FFF9C4', font=('System',17),).place(x=120, y=15)
    Label(rooms, text='Email: ' + str(ws['G'+str(x)].value), bg='#FFF9C4', font=('System',5),).place(x=125, y=50)
    Label(rooms, text='Phone: ' + str(ws['F'+str(x)].value), bg='#FFF9C4', font=('System',5),).place(x=125, y=70)
    Label(rooms, text= str(ws['D'+str(x)].value), bg='#FFF9C4', font=('System',5),).place(x=420, y=10)
    Label(rooms, text= str(ws['E'+str(x)].value), bg='#FFF9C4', font=('System',5),).place(x=420, y=30)
    Label(rooms, text= str(ws['B'+str(x)].value), bg='#FFF9C4', font=('System',2),).place(x=120, y=0)  
    detail = Frame(rooms, width=600, height=700, bg='#FFF176')
    detail.place(x=0, y=100)
    Label(detail, text='Monthly Deposite:   ' + str(ws['C'+str(x)].value), bg='#FFF176', font=('System',17),).place(x=20, y=30) 
    Label(detail, text='Missed Payments:   ' + str(ws['H'+str(x)].value), bg='#FFF176', font=('System',17),).place(x=20, y=100) 
    Label(detail, text='Damages:   ' + str(ws['I'+str(x)].value), bg='#FFF176', font=('System',17),).place(x=20, y=170) 
    Label(detail, text='Issue:   ' + str(ws['J'+str(x)].value), bg='#FFF176', font=('System',17),).place(x=20, y=240) 
    Label(detail, text='Rental History:', bg='#FFF176', font=('System',17),).place(x=210, y=350) 
    Label(detail, text=str(ws['K'+str(x)].value), bg='#FFF176', font=('System',17),).place(x=20, y=400) 
    Evict = Button(detail, text='EVICT', font=('System',17), command=lambda:kick(x))
    Evict.place(x=250, y=550)
  else:
    Label(rooms, text='Name:', bg='#FFF9C4', font=('System',17)).place(x=0, y=150)
    T1= Entry(rooms, width=18, font=('System',17) )
    T1.place(x=260, y=150)
    Label(rooms, text='SSN:', bg='#FFF9C4', font=('System',17)).place(x=0, y=200)  
    T2= Entry(rooms, width=18, font=('System',17) )
    T2.place(x=260, y=200)
    Label(rooms, text='Monthly Deposite:', bg='#FFF9C4', font=('System',17)).place(x=0, y=250) 
    T3= Entry(rooms, width=18, font=('System',17) )
    T3.place(x=260, y=250)
    Label(rooms, text='Barangay:', bg='#FFF9C4', font=('System',17)).place(x=0, y=300) 
    T4= Entry(rooms, width=18, font=('System',17) )
    T4.place(x=260, y=300)
    Label(rooms, text='City/Province:', bg='#FFF9C4', font=('System',17)).place(x=0, y=350) 
    T5= Entry(rooms, width=18, font=('System',17) )
    T5.place(x=260, y=350)
    Label(rooms, text='Phone number:', bg='#FFF9C4', font=('System',17)).place(x=0, y=400) 
    T6= Entry(rooms, width=18, font=('System',17) )
    T6.place(x=260, y=400)
    Label(rooms, text='Email:', bg='#FFF9C4', font=('System',17)).place(x=0, y=450) 
    T7= Entry(rooms, width=18, font=('System',17) )
    T7.place(x=260, y=450)
    Label(rooms, text='Missed payments:', bg='#FFF9C4', font=('System',17)).place(x=0, y=500) 
    T8= Entry(rooms, width=18, font=('System',17) )
    T8.place(x=260, y=500)
    Label(rooms, text='Damages:', bg='#FFF9C4', font=('System',17)).place(x=0, y=550) 
    T9= Entry(rooms, width=18, font=('System',17) )
    T9.place(x=260, y=550)
    Label(rooms, text='Complaints:', bg='#FFF9C4', font=('System',17)).place(x=0, y=600) 
    T10= Entry(rooms, width=18, font=('System',17) )
    T10.place(x=260, y=600)
    Label(rooms, text='Rental History:', bg='#FFF9C4', font=('System',17)).place(x=0, y=650) 
    T11= Entry(rooms, width=18, font=('System',17) )
    T11.place(x=260, y=650)
    Apply = Button(rooms, text='Enter', font=('System',17), command=lambda:apply(x))
    Apply.place(x=250, y=720)

Building = ctk.CTkScrollableFrame(cmd, height=535, width=1180, corner_radius=0)
Floor = Frame(Building, width=1200, height=750, bg='#FFF59D')
Room1 = Button(Floor, text='Occupied', width= 250, fg='Red',bg='#FFF571', activebackground='#FFE664', activeforeground='red', font=('System',17), height=200, cursor='hand2', image=Occupied, compound=BOTTOM, command=lambda:occupy(1))
Room2 = Button(Floor, text='Occupied', width= 250, fg='Red',bg='#FFF571', activebackground='#FFE664', activeforeground='red', font=('System',17), height=200, cursor='hand2', image=Occupied, compound=BOTTOM, command=lambda:occupy(2))
Room3 = Button(Floor, text='Vacant', width= 250, fg='Black',bg='#FFF571', activebackground='#FFE664', activeforeground='black', font=('System',17), height=200, cursor='hand2', image=Vacant, compound=BOTTOM, command=lambda:occupy(3))
Room4 = Button(Floor, text='Occupied', width= 250, fg='Red',bg='#FFF571', activebackground='#FFE664', activeforeground='red', font=('System',17), height=200, cursor='hand2', image=Occupied, compound=BOTTOM, command=lambda:occupy(4))
Room5 = Button(Floor, text='Vacant', width= 250, fg='Black',bg='#FFF571', activebackground='#FFE664', activeforeground='black', font=('System',17), height=200, cursor='hand2', image=Vacant, compound=BOTTOM, command=lambda:occupy(5))
Room6 = Button(Floor, text='Occupied', width= 250, fg='Red',bg='#FFF571', activebackground='#FFE664', activeforeground='red', font=('System',17), height=200, cursor='hand2', image=Occupied, compound=BOTTOM, command=lambda:occupy(6))

#Employee
def person(cell):
  global ws, num1, num2, num3, num4, num5, num6, num7, num8, num9, mm1, mm2, mm3, mm4, mm5, mm6, mm7, mm9, store
  ws = exl['Employee']
  store = cell
  Update.place(x=650, y=450)
  Del.place(x=650, y= 550)
  Save.place_forget()
  for cells in range(2, ws.max_row + 1):
    if cell ==  ws['A'+str(cells)].value:
      reset()
      mm1.insert(0,ws['A'+str(cells)].value)
      mm2.insert(0,ws['B'+str(cells)].value)
      mm3.insert(0,ws['C'+str(cells)].value)
      mm4.insert(0,ws['D'+str(cells)].value)
      mm5.insert(0,ws['E'+str(cells)].value)
      mm6.insert(0,ws['F'+str(cells)].value)
      mm7.insert(0,ws['G'+str(cells)].value)
      if (ws['H'+str(cells)].value) == ("Male"):
        Male.select()
        Female.deselect()
      elif (ws['H'+str(cells)].value) == ("Female"):
        Male.deselect()
        Female.select()
      mm9.insert("1.0",ws['I'+str(cells)].value)
      collect()
      break      

def record():
  reset()
  Save.place(x=650, y=460)
  Update.place_forget()
  Del.place_forget()

def collect():
    global num1, num2, num3, num4, num5, num6, num7, num8, num9 #to allow the variable be used everywhere
    num1= mm1.get()
    num2= mm2.get()
    num3= mm3.get()
    num4= mm4.get()
    num5= mm5.get()
    num6= mm6.get()
    num7= mm7.get()
    num8= Gender.get()
    num9= mm9.get("1.0","end-1c") #Text

def save():
  found = True
  collect()
  if num1==("") or num2==("") or num3==("") or num4==("") or num5 ==("") or num6==("") or num7==("") or num8 ==("") or num9 ==(""):
        messagebox.showwarning("Incomplete Data", 'Complete the form')
  else:
      for cells in range(2, ws.max_row + 1):
        if num1 == ws['A'+str(cells)].value:
          found = True
          break
        else:
           found = False
      if found == False: 
          Nrow= str(ws.max_row + 1)
          ws['A'+ Nrow]= num1
          ws['B'+ Nrow]= num2
          ws['C'+ Nrow]= num3
          ws['D'+ Nrow]= num4
          ws['E'+ Nrow]= num5
          ws['F'+ Nrow]= num6
          ws['G'+ Nrow]= num7
          ws['H'+ Nrow]= num8   
          ws['I'+ Nrow]= num9
          exl.save('AMS.xlsx')
          messagebox.showinfo("Success", "employee data was Updated")
          reset()
          Employ()
          
      else:
          found = True

def update():
  ws = exl['Employee']
  found = True
  if num1==("") or num2==("") or num3==("") or num4==("") or num5 ==("") or num6==("") or num7==("") or num8 ==("") or num9 ==(""):
        messagebox.showwarning("Incomplete Data", 'Complete the form')
  else:  
    collect()
    for cells in range(2, ws.max_row + 1):
      if store == ws['A'+str(cells)].value:
        found = False
        row = cells
        break
      else:
        found = True
    if found == False:
      ws['A'+ str(row)]= num1
      ws['B'+ str(row)]= num2
      ws['C'+ str(row)]= num3
      ws['D'+ str(row)]= num4
      ws['E'+ str(row)]= num5
      ws['F'+ str(row)]= num6
      ws['G'+ str(row)]= num7
      ws['H'+ str(row)]= num8  
      ws['I'+ str(row)]= num9
      exl.save('AMS.xlsx')
      messagebox.showinfo("Success!","Your data has been added to the employee details.")
      reset()
      Employ()

def remove():
    ws = exl['Employee']
    for cells in range(2, ws.max_row + 1):
      if num1 == ws['A'+str(cells)].value:
          row= cells
          found = True
          break
      else:
          found = False
    if found == True:
        ws.delete_rows(row)
        dead = name
        exl.save('AMS.xlsx')        
        messagebox.showwarning('BOOOOM!', "Data was permanently removed")
        dead.destroy()
        reset()
        Employ()
    else:
        messagebox.showerror("Data not found", "Reference number does not match any Employee")

def reset():
  mm1.delete(0,END)
  mm2.delete(0,END)
  mm3.delete(0,END)
  mm4.delete(0,END)
  mm5.delete(0,END)
  mm6.delete(0,END)
  mm7.delete(0,END)
  Male.deselect()
  Female.deselect()
  mm9.delete("1.0","end-1c")

def Employ():
  global name, mm1, mm2, mm3, mm4, mm5, mm6, mm7, mm8, mm9, ws
  ws = exl['Employee']
  for cells in range(2, ws.max_row + 1):
    if ws['A'+str(cells)].value != '':
      unit = int(cells-1)
      name = ws['A'+str(cells)].value
      name = Button(Selection, height=5, width=11, text=str(name), borderwidth=0, command=lambda cell=name:person(cell))
      new_height= 0+int(85 * int(unit - 1))
      name.place(x=0, y=new_height)
      AR.place(x=0, y=int(new_height+85))
      if 650-new_height > 0:
        Selection.configure(height=650)
      else:
        Selection.configure(height=650+int(new_height))
    else:
      break
  Save.place_forget()
  Update.place(x=650, y=460)
  Del.place(x=650, y= 550)
  Reset.place(x=850, y=460)
  mm1.place(x=300, y=50)
  mm2.place(x=750, y=50)
  mm3.place(x=300, y=150)
  mm4.place(x=750, y=150)
  mm5.place(x=300, y=250)
  mm6.place(x=750, y=250)
  mm7.place(x=300, y=350)
  Male.place(x=750, y=340)
  Female.place(x=880, y=340)
  mm9.place(x=150, y=450)

Frame1 = ctk.CTkScrollableFrame(cmd, height=535, width=1180, corner_radius=0)
Containter = Frame(Frame1, width=1200, height=650, bg='#C0AA40' )
Selection = Frame(Frame1, width=85, height=535)
Selection.pack(side='left')
Update = Button(Containter, text="Update", font=('System',17), width= 7, command=lambda:update())
Save = Button(Containter, text="Add", font=('System',17), width= 7, command=lambda :save())
Reset= Button(Containter, text="Reset", font=('System',17), width= 7,command=lambda:reset())
Del= Button(Containter, text="Delete", font=('System',17), width= 7, command=lambda:remove())
AR= Button(Selection, text="+", font=('System',17), width= 5, height=2, borderwidth=0.5, command=lambda:record())
mm1 = Entry(Containter, width=15, font=('Helvetica',20), fg='black')
mm2 = Entry(Containter, width=15, font=('Helvetica',20), fg='black')
mm3 = Entry(Containter, width=15, font=('Helvetica',20), fg='black')
mm4 = Entry(Containter, width=15, font=('Helvetica',20), fg='black')
mm5 = Entry(Containter, width=15, font=('Helvetica',20), fg='black')
mm6 = Entry(Containter, width=15, font=('Helvetica',20), fg='black')
mm7 = Entry(Containter, width=15, font=('Helvetica',20), fg="black")
Gender= StringVar()
Male= Radiobutton(Containter, width= 8,text="Male", font=('',15), indicatoron=FALSE, variable=Gender, value= 'Male')
Male.deselect()
Female= Radiobutton(Containter, width= 8, text= "Female", font=('',15), indicatoron=FALSE, variable=Gender, value='Female')
Female.deselect()
mm9 = Text(Containter, width=25, height=5, font=('Helvetica',20), fg='black')

#Menu
Category= Frame(cmd, height=65, width=1200, bg='#22262D')
Category.grid(row=0, column=0)
#Home is bugged
Home= Button(Category, text='Home', bg='#22262D', state='disabled', font=('Helvetica', 26), borderwidth=0, foreground='White', activebackground='#C0AA40', command=lambda:tab1())
Room= Button(Category, text='Apartment', bg='#22262D', font=('Helvetica', 26), borderwidth=0, foreground='White', activebackground='#C0AA40', command=lambda:tab2())
Employee = Button(Category, text='Employee', bg='#22262D', font=('Helvetica', 26), borderwidth=0, foreground='White', activebackground='#C0AA40', command=lambda:tab5())
Exit= Button(Category, text='Logout', compound='right', image=Logout, bg='#22262D', font=('Helvetica', 26), borderwidth=0, foreground='White', activebackground='#C0AA40', command=lambda:exit())

Home.place(x=100, y=0)
Room.place(x=350, y=0)
Employee.place(x=700, y=0)
Exit.place(x=1000, y=0)


tab1()
cmd.mainloop()