from tkinter import *
import tkinter as tk
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk
from tkinter import messagebox

#Excel
exl = Workbook()
exl = load_workbook('AMS.xlsx')

#main window
lgs = tk.Tk()
#computations
heights= 800
widths= 1200
screenx= lgs.winfo_screenwidth()
screeny= lgs.winfo_screenheight()
leftscrn= int(screenx/2 - widths/2)
topscrn= 0
topscrn2= int(screeny/2 - (heights-200)/2)
#configurations
lgs.geometry(f'{widths}x{heights}+{leftscrn}+{topscrn}')
lgs.title('My appartment manager')
lgs.resizable(False, False)
lgs.iconbitmap('TetraICO.ico')

#images
walls = Image.open('LoginInterface.png')
wall= ImageTk.PhotoImage(walls)
hide= Image.open('unhide.png')
hide= hide.resize((26,21), Image.ADAPTIVE)
unhide= ImageTk.PhotoImage(hide)

#functions
def reg(): # Switch Login/Register Interface
    global Regis, Login
    if Regis : #Registeration interface
        Login = False
        Username.delete(0,'end')
        Password.delete(0,'end')
        Password.configure(show='', width= 35)
        show_pass.place_forget()
        quote1.place_forget()
        quote2.place(x=100, y= 396)
        Register.configure(text='Login')
        Register.place(x=300)
        Start.place_forget()
        Reg.place(x= 56, y= 457)
        Regis = False
    else: #Login interface
        Login = True
        Username.delete(0,'end')
        Password.delete(0,'end')
        Password.configure(show='•', width=32)
        show_pass.place(x=362, y=0)
        quote2.place_forget()
        quote1.place(x=160, y= 396)
        Register.configure(text='Register')
        Register.place(x=260)
        Reg.place_forget()
        Start.place(x= 56, y= 457)
        Regis = True

def show(): #Show password Button
    global visible
    if visible:
        Password.configure(show='•')
        visible= False
    else:
        Password.configure(show='')
        visible= True

def new(): #Login/Register Button
    ws= exl['Accounts']
    global found, Nrow, num1, num2, Login
    found= False
    data()
    if Login == False: #Registration interface
        ws= exl['Accounts']
        if Username.get() == "":
            messagebox.showerror('No username','Enter your Username')
        elif Password.get() == "":
            messagebox.showerror('No Password','Enter a Password')
        else:
            for cells in range(2, ws.max_row + 1):
                if num1 == ws['A'+str(cells)].value:
                    found = True
                    break
                else:
                    found = False
            if found:
                messagebox.showerror('Name Taken','Username was already taken.')
                Username.delete(0,'end')
            else:
                Nrow= str(ws.max_row + 1)
                ws['A'+ Nrow]= num1
                ws['B'+ Nrow]= num2
                messagebox.showinfo('Account has been Created','You are done creating account')
                Login = True
                reg()
    else: #Login Interface
        exist = False
        if Username.get() == "":
            messagebox.showerror('No username','Enter your Username')
        elif Password.get() == "":
            messagebox.showerror('No Password','Enter a Password')
        else:
            ws= exl['Accounts']
            for rows in range(2, ws.max_row + 1):
                if (num1 == ws['A'+str(rows)].value) and (num2 == ws['B'+str(rows)].value):
                    exist = True
                    break
                else:
                    exist = False
            if exist == True:
                for rows in range(2, ws.max_row + 1):
                    if (num1 == ws['A'+str(rows)].value):
                        row= rows
                        User= ws['A'+str(row)].value
                messagebox.showinfo("Login Success", f'Welcome {User}')
                lgs.withdraw()
            else:
                messagebox.showerror('Account doesnt Exist', 'Username or Password Incorrect')
    exl.save('AMS.xlsx')

def data():
    global num1, num2
    num1= Username.get()
    num2= Password.get()

bgm= Canvas(lgs, background='Pink', width=1200, height=800)
bgm.place(x= 0, y= 0, relwidth=1, relheight=1)
Label(bgm, image=wall).place(x=0, y=0, relheight=1, relwidth= 1)

outline1= LabelFrame(lgs, text='Username', width=400, height=45, font=('COPPERPLATE GOTHIC LIGHT', 10), bg='#704A27')
outline1.place(x=61, y= 234)
Username= Entry(outline1, width=35, font= 'Helvetica 15')
Username.place(x=4, y=0)
outline2= LabelFrame(lgs, text='Password', width=400, height=45, font=('COPPERPLATE GOTHIC LIGHT', 10), bg='#835627')
outline2.place(x=61, y= 307)
Password= Entry(outline2, width=32, font= 'Helvetica 15', show='•')
Password.place(x=4, y=0)
quote= Label(lgs, text='“Experience the comfort of home away from home”', font=('COPPERPLATE GOTHIC LIGHT', 10),bg='#FFDE59')
quote.place(x=69, y= 198)
quote1= Label(lgs, text='New user?', font=('COPPERPLATE GOTHIC LIGHT', 10),bg='#FFDE59')
quote1.place(x=160, y= 396)
quote2= Label(lgs, text='Already have an account?', font=('COPPERPLATE GOTHIC LIGHT', 10),bg='#FFDE59')
Register= Label(lgs, text='Register', font=('COPPERPLATE GOTHIC LIGHT', 10,'underline'),bg='#FFDE59', cursor='hand2')
Register.place(x=260, y= 396)
Register.bind('<Button-1>', lambda event: reg())

Regis = True
Login = True
visible= False
show_pass= Button(outline2, image=unhide, cursor='hand2', command=show)
show_pass.place(x=362, y=0)
Reg= Button(lgs, width=19, height= 1, bg='#C0AA40', text='Register', font=('Helvetica',26),fg="White", activebackground="#867314", activeforeground="White", cursor='hand2', command=new)
Start= Button(lgs, width=19, height= 1, bg='#C0AA40', text='Login', font=('Helvetica',26),fg="White", activebackground="#867314", activeforeground="White", cursor='hand2', command=new)
Start.place(x= 56, y= 457)

lgs.mainloop()